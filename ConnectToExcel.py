import openpyxl as opx
import os
from tabulate import tabulate

class ExcelFile:
    
    def __init__(self, name):
        # Load the excel file, and the specific sheet
        self.wb = opx.load_workbook(name)

        # make a Queries folder if it doesn't exist
        if not os.path.exists('Queries'):
            os.mkdir("Queries")

    # This list will hold the data from the spreadsheet
    # Ignoring the first row since those are the column names
    def CreateTableDataFromWorksheet(self, ws):
        TableData = []
        for i in range(ws.max_row-1):
            TableData.append([])
            
        for i in range(2, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                CellValue = ws.cell(row = i, column = j).value
                TableData[i-2].append(CellValue)
        return TableData

    # Headers from the spreadsheet, i.e. the first row
    def CreateHeadersFromWorksheet(self, ws):
        Headers = []
        for i in range(1, ws.max_column + 1):
            CellValue = ws.cell(row = 1, column = i).value
            Headers.append(CellValue)
        return Headers

    def GetGroupsFromRoleAndLocation(self, role, place):
        
        # initalize a random sheet so Python knows sheet's type
        sheet = self.wb.active

        # Load worksheet associated with location
        if place == "Summit" or place == "Warren" or place == "SanDiego":
            sheet = self.wb["Summit_Warren_SanDiego_Groups"]
        elif place == "Phoenix":
            sheet = self.wb["Phoenix_Groups"]
        elif place == "Juno":
            sheet = self.wb["JunoUsers_Groups"]
        elif place == "Boudry" or place == "Couvet":
            sheet = self.wb["Boudry_Couvet_Groups"]
        elif place == "S12":
            sheet = self.wb["SummitS12CarTUsers_Groups"]

        # Data in table not including column names
        TableData = self.CreateTableDataFromWorksheet(sheet)

        # Column names
        Headers   = self.CreateHeadersFromWorksheet(sheet)

        # Empty list for us to return groups of interest
        Groups = []

        # Get index of specified role i.e. column column
        RoleIdx = Headers.index(role)

        # Get index of Group i.e. column index
        GroupIdx = Headers.index("Groups")

        # Find all groups that specified role is assocaited with
        for Row in TableData:
            if Row[RoleIdx] == "Yes":
                Groups.append(Row[GroupIdx])

        return Groups

    # Print out the table from the worksheet
    def PrintTableFromWorksheet(self, ws):
        TableData = self.CreateTableDataFromWorksheet(ws)
        Headers = self.CreateHeadersFromWorksheet(ws)
        print(tabulate(TableData, headers=Headers, tablefmt='orgtbl'))

    def PrintAllTables(self):
        for sheet in self.wb.worksheets:
            self.PrintTableFromWorksheet(sheet)
            print("")

    def CreateTableQuery(self, ws):
        # Insert portion of query
        file = open("Queries/" + ws.title + "_TableCreation.txt", 'w')
        string = "CREATE TABLE IF NOT EXISTS " + ws.title
        file.write(string)
        string = ""

        # Header field portion of query        
        Headers = self.CreateHeadersFromWorksheet(ws)
        fields = "\n(\n"
        for i in range(len(Headers)):
            fields += "\t" + Headers[i]
            Cell = ws.cell(row = 2, column = i + 1).value
            if isinstance(Cell, str):
                fields += " VARCHAR(100)"
            if isinstance(Cell, int):
                fields += " INT"
            if isinstance(Cell, float):
                fields += " DEC(4,2)" 
            if (i != len(Headers) - 1):
                fields += ", "
            fields += "\n"
        string += fields + ");"
        file.write(string)
        string = ""
        return ws.title + "_TableCreation.txt"

    def CreateInsertQuery(self, ws):
        # Insert portion of query
        file = open("Queries/" + ws.title + "_Insertion.txt", 'w')
        string = "INSERT INTO " + ws.title
        file.write(string)
        string = ""

        # Header field portion of query        
        Headers = self.CreateHeadersFromWorksheet(ws)
        fields = "\n("
        for i in range(len(Headers)):
            fields += Headers[i]
            if (i != len(Headers) - 1):
                fields += ", "
        string += fields + ")"
        file.write(string)
        string = ""

        # Beginning of values portion of query
        file.write("\nVALUES\n")

        # Values portion of query
        TableData = self.CreateTableDataFromWorksheet(ws)
        values = ""
        for i in range(len(TableData)):
            value = "("
            for j in range(len(TableData[i])):
                if isinstance(TableData[i][j], str):
                    value += "'"
                value += str(TableData[i][j])
                if isinstance(TableData[i][j], str):
                    value += "'"
                if (j != len(TableData[i]) - 1):
                    value += ", "
            value += ")"
            if (i != len(TableData) - 1):
                    value += ",\n"
            values += value
        string += values
        string += ";"
        file.write(string)
        return ws.title + "_Insertion.txt"

    def GetGroupMemberList(self, ws, db):
        # Headers are column names from excelfile
        Headers = self.CreateHeadersFromWorksheet(ws)

        # Table data is everything but the column names
        TableData = self.CreateTableDataFromWorksheet(ws)

        # Get indices of Column names
        UnameIdx = Headers.index("Username")
        SiteIdx = Headers.index("Site")
        RoleIdx = Headers.index("Role")
        AddRemIdx = Headers.index("AddRemove")

        # Get SecID from Username Query
        SecID_query = "SELECT SecID FROM UserNames Where Username = "

        # Get GroupID from GroupName Query
        GroupID_query = "SELECT GroupID FROM GroupTable WHERE GroupName = "
        
        # This field will contain SecID and GroupID
        GroupMembers = []

        # Get the SecIDs and GroupIDs
        for Row in TableData:
            db.RunQuery(SecID_query + "'" + Row[UnameIdx] + "'")
            SecID = db.cur.fetchone()[0]
            # Check if we are making an insert i.e. row in AddRemove column says "Add"
            if Row[AddRemIdx] == "Add":
                roles = Row[RoleIdx].split(", ") # allow for multiple roles
                for role in roles:
                    groups = self.GetGroupsFromRoleAndLocation(role, Row[SiteIdx])
                    for group in groups:
                        db.RunQuery(GroupID_query + "'" + group + "'")
                        GroupID = db.cur.fetchone()[0]
                        GroupMembers.append([SecID, GroupID])
        return GroupMembers
                
    def CreateGroupMemberInsertQuery(self, ws, db):
        file = open("Queries/Insertion.txt", 'w')
        string = "INSERT INTO GroupMembers\n(SecID, GroupID)"
        string += "\nVALUES\n"
    
        # Get the group members i.e. the list that contains SecID and GroupID
        GroupMembers = self.GetGroupMemberList(ws, db)

        # Part of query that adds values as tuples
        values = ""
        for i in range(len(GroupMembers)):
            value = "("
            for j in range(len(GroupMembers[i])):
                value += str(GroupMembers[i][j])
                if (j != len(GroupMembers[i]) - 1):
                    value += ", "
            value += ")"
            if (i != len(GroupMembers) - 1):
                    value += ",\n"
            values += value
        string += values
        string += ";"
        if len(GroupMembers) != 0:
            file.write(string)
        else:
            file.write("blank")
        file.close()
        
    def CreateGroupMemberDeletionQuery(self, ws, db):
        # Create the file for deletion query
        file = open("Queries/GroupMemberDeletion.txt", 'w')
        string = "DELETE FROM GroupMembers WHERE SecID = "
        
        # Headers are column names from excelfile
        Headers = self.CreateHeadersFromWorksheet(ws)

        # Table data is everything but the column names
        TableData = self.CreateTableDataFromWorksheet(ws)

        # Column index of column name username
        UnameIdx = Headers.index("Username")
        AddRemIdx = Headers.index("AddRemove")

        # Get SecID from Username Query
        SecID_query = "SELECT SecID FROM UserNames Where Username = "

        # Remove query based on SecID
        AtLeastOneDeletion = False
        for i in range(len(TableData)):
            if TableData[i][AddRemIdx] == "Remove":
                AtLeastOneDeletion = True
                db.RunQuery(SecID_query + "'" + TableData[i][UnameIdx] + "'")
                SecID = db.cur.fetchone()[0]
                if i != 0:
                    string += "\nOR SecID = "
                string += str(SecID)
        string += ";"
        if AtLeastOneDeletion:
            file.write(string)
        else:
            file.write("blank")

    def CreateUsernameDeletionQuery(self, ws):
        # Create the file for deletion query
        file = open("Queries/UsernameDeletion.txt", 'w')
        string = "DELETE FROM usernames WHERE username = "
        
        # Headers are column names from excelfile
        Headers = self.CreateHeadersFromWorksheet(ws)

        # Table data is everything but the column names
        TableData = self.CreateTableDataFromWorksheet(ws)

        # Column index of column name username
        UnameIdx = Headers.index("Username")
        AddRemIdx = Headers.index("AddRemove")

        # Remove query based on SecID
        AtLeastOneDeletion = False
        for i in range(len(TableData)):
            if TableData[i][AddRemIdx] == "Remove":
                AtLeastOneDeletion = True
                if i != 0:
                    string += "\nOR Username = "
                string += "'" + TableData[i][UnameIdx] + "'"
        string += ";"
        if AtLeastOneDeletion:
            file.write(string)
        else:
            file.write("blank")

    def EliminateDuplicates(self, fileName):
        # Open file for reading
        readFile = open(fileName, "r")

        # Get file line by line
        lines = readFile.readlines()

        # Eliminate duplicates
        NoDupes = list(dict.fromkeys(lines))

        # Close reading file and open write file
        readFile.close()
        writeFile = open(fileName, "w")

        # Write back to the file
        for line in NoDupes:
            writeFile.write(line)

        # Close write file
        writeFile.close()

